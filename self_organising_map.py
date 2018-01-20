#!/usr/bin/env python
# Tested on Ubuntu
import numpy as np
import os

from pylab import *
from openpyxl import *

def fromdistance(fn, shape, center=None, dtype=float):
    def distance(*args):
        d = 0
        for i in range(len(shape)):
            d += ((args[i]-center[i])/float(max(1,shape[i]-1)))**2
        return np.sqrt(d)/np.sqrt(len(shape))
    if center == None:
        center = np.array(list(shape))//2
    return fn(np.fromfunction(distance,shape,dtype=dtype))

def Gaussian(shape,center,sigma=0.5):
    ''' '''
    def g(x):
        return np.exp(-x**2/sigma**2)
    return fromdistance(g,shape,center)

class SOM:
    ''' Self-organizing map '''

    def __init__(self, *args):
        ''' Initialize som '''
        self.codebook = np.zeros(args)
        self.reset()

    def reset(self):
        ''' Reset weights '''
        self.codebook = np.random.random(self.codebook.shape)

    def learn(self, samples, epochs=10000, sigma=(10, 0.001), lrate=(0.5,0.005)):
        ''' Learn samples '''
        sigma_i, sigma_f = sigma
        lrate_i, lrate_f = lrate

        for i in range(epochs):
            # Adjust learning rate and neighborhood
            t = i/float(epochs)
            lrate = lrate_i*(lrate_f/float(lrate_i))**t
            sigma = sigma_i*(sigma_f/float(sigma_i))**t

            # Get random sample
            index = np.random.randint(0,samples.shape[0])
            data = samples[index]

            # Get index of nearest node (minimum distance)
            D = ((self.codebook-data)**2).sum(axis=-1)
            winner = np.unravel_index(np.argmin(D), D.shape)

            # Generate a Gaussian centered on winner
            G = Gaussian(D.shape, winner, sigma)
            G = np.nan_to_num(G)

            # Move nodes towards sample according to Gaussian
            delta = self.codebook-data
            for i in range(self.codebook.shape[-1]):
                self.codebook[...,i] -= lrate * G * delta[...,i]


# -----------------------------------------------------------------------------
if __name__ == '__main__':
    import matplotlib
    import matplotlib.pyplot as plt
    try:    from voronoi import voronoi
    except: voronoi = None

    def learn(network, samples, epochs=25000, sigma=(10, 0.01), lrate=(0.5,0.005)):
        network.learn(samples, epochs)
        fig = plt.figure(figsize=(5,5))
        axes = fig.add_subplot(1,1,1)
        # Draw samples
        x,y = samples[:,0], samples[:,1]
        plt.scatter(x, y, s=10, color='r', marker="+")
        # Draw network
        x,y = network.codebook[...,0], network.codebook[...,1]
        if len(network.codebook.shape) > 2:
            for i in range(network.codebook.shape[0]):
                plt.plot (x[i,:], y[i,:], 'k', alpha=0.85, lw=1.5, zorder=2)
            for i in range(network.codebook.shape[1]):
                plt.plot (x[:,i], y[:,i], 'k', alpha=0.85, lw=1.5, zorder=2)
        else:
            plt.plot (x, y, 'k', alpha=0.85, lw=1.5, zorder=2)
        plt.scatter (x, y, s=50, c='w', edgecolors='k', zorder=3)
        if voronoi is not None:
            segments = voronoi(x.ravel(),y.ravel())
            lines = matplotlib.collections.LineCollection(segments, color='0.65')
            axes.add_collection(lines)
        plt.axis([0,1,0,1])
        plt.xticks([]), plt.yticks([])
        plt.show()
#Fonction d importation des donnees
def import_data():
    nb_term=[]
    data=[]
    #Importation des donnees dans le programme
    wb = load_workbook(filename = 'dataframe_INRIA.xlsx')
    sheet=wb.active
    for i in range(2,4047): #4044 lignes
            a1=sheet.cell(row=i, column=1)
            a2=sheet.cell(row=i, column=2)
            nb_term.append(int(a1.value))
            data.append(float(a2.value))
    samples = np.zeros((len(data),2))
    for i in range(len(data)):
        samples[i][1]=data[i]
        samples[i][0]=nb_term[i]
    return samples
#Fonction qui retourne les valeurs d une serie "num"
def getSerie(samples,num):
    values=[]
    for i in range(len(samples)):
        if (samples[i][0]==num):
            values.append(samples[i][1])
    return values
#Permet d obtenir les taux d accroissement entre chaque terme d une suite num
def getGrowthRate(samples,num):
    serie1=getSerie(samples,num)
    growthRate=[]
    if(len(serie1)==1):
        growthRate.append(1.0)
    else:
        for i in range(len(serie1)-1):
            growthRate.append(float(serie1[i+1]-serie1[i]))
    return growthRate

def getStdDeviation(samples,num):
    serie=getGrowthRate(samples,i)
    acc=0
    for j in range(len(serie)):
        acc+=abs(serie[j]-mean(getGrowthRate(samples,i)))
    acc=acc/len(serie)
    acc=sqrt(acc)
    return(acc)

# --- main ---- #
samples=import_data()
data_kohonen = np.zeros((81,2))
x=[]
y=[]

print "Performing data preparation"
for i in range(81):
    x.append(float(len(getSerie(samples,i))))
    #print " -- Length : ", len(getSerie(samples,i))
    y.append(float(getStdDeviation(samples,i)))
    #print " -- Standart Deviation : ",getStdDeviation(samples,i)
    data_kohonen[i][0]=x[i]
    data_kohonen[i][1]=y[i]

print "Done"
'''
print "See the result :"
scatter(x,y,marker='+')
show()
os.system("clear")
'''
print "Normalizing data"
max_sigma=data_kohonen[0][1]
max_len=data_kohonen[0][0]
for i in range(81):
    if(max_sigma<data_kohonen[i][1]):
        max_sigma=data_kohonen[i][1]
    if(max_len<data_kohonen[i][0]):
        max_len=data_kohonen[i][0]
for i in range(81):
    data_kohonen[i][0]=(data_kohonen[i][0]*0.9)/max_len
    data_kohonen[i][1]=(data_kohonen[i][1]*0.9)/max_sigma
print "Done"
print "Applying Kohonen algorithm"
kohonen=SOM(4,4,2)
learn(kohonen,data_kohonen)
print "Done"
print "END"
